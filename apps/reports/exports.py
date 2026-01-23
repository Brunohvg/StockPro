"""
Product Exporter - Multi-format export for products and movements
"""
import csv
import io
import json
from datetime import datetime
from decimal import Decimal

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from apps.products.models import Product, ProductType


class DecimalEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        if isinstance(obj, datetime):
            return obj.isoformat()
        return super().default(obj)


class ProductExporter:
    """Export products and variants to various formats"""

    def __init__(self, tenant):
        self.tenant = tenant

    def get_products(self, include_inactive=False):
        """Fetch all products with variants"""
        qs = Product.objects.filter(tenant=self.tenant).select_related(
            'category', 'brand'
        ).prefetch_related(
            'variants', 'variants__attribute_values', 'variants__attribute_values__attribute_type'
        )
        if not include_inactive:
            qs = qs.filter(is_active=True)
        return qs.order_by('name')

    def _simple_row(self, product):
        """Generate CSV row for simple product"""
        return {
            'sku': product.sku,
            'name': product.name,
            'type': 'SIMPLE',
            'category': product.category.name if product.category else '',
            'brand': product.brand.name if product.brand else '',
            'uom': product.uom,
            'stock': product.current_stock,
            'minimum_stock': product.minimum_stock,
            'cost': float(product.avg_unit_cost) if product.avg_unit_cost else 0,
            'barcode': product.barcode or '',
        }

    def _parent_row(self, product):
        """Generate CSV row for variable product (parent)"""
        return {
            'sku': product.sku,
            'name': product.name,
            'type': 'VARIABLE',
            'category': product.category.name if product.category else '',
            'brand': product.brand.name if product.brand else '',
            'uom': product.uom,
            'stock': '',  # Stock is on variants
            'minimum_stock': '',
            'cost': '',
            'barcode': '',
        }

    def _variant_row(self, variant, attr_columns):
        """Generate CSV row for variant"""
        row = {
            'sku': variant.sku,
            'name': variant.name or '',
            'type': f'VARIANT:{variant.product.sku}',
            'category': '',
            'brand': '',
            'uom': variant.product.uom,
            'stock': variant.current_stock,
            'minimum_stock': variant.minimum_stock,
            'cost': float(variant.avg_unit_cost) if variant.avg_unit_cost else 0,
            'barcode': variant.barcode or '',
        }

        # Add attribute values
        for attr in variant.attribute_values.all():
            col_name = f'attr_{attr.attribute_type.name.lower()}'
            row[col_name] = attr.value

        # Ensure all attr columns exist
        for col in attr_columns:
            if col not in row:
                row[col] = ''

        return row

    def _get_all_attr_columns(self, products):
        """Collect all unique attribute column names"""
        attr_cols = set()
        for product in products:
            if product.is_variable:
                for variant in product.variants.all():
                    for attr_val in variant.attribute_values.all():
                        attr_cols.add(f'attr_{attr_val.attribute_type.name.lower()}')
        return sorted(attr_cols)

    def export_csv(self, include_variants=True, include_inactive=False):
        """Export products to CSV format"""
        products = self.get_products(include_inactive)
        attr_columns = self._get_all_attr_columns(products) if include_variants else []

        rows = []
        for product in products:
            if product.product_type == ProductType.SIMPLE:
                row = self._simple_row(product)
                for col in attr_columns:
                    row[col] = ''
                rows.append(row)
            else:
                row = self._parent_row(product)
                for col in attr_columns:
                    row[col] = ''
                rows.append(row)

                if include_variants:
                    for variant in product.variants.filter(is_active=True):
                        rows.append(self._variant_row(variant, attr_columns))

        # Build CSV
        output = io.StringIO()
        base_fieldnames = ['sku', 'name', 'type', 'category', 'brand', 'uom', 'stock', 'minimum_stock', 'cost', 'barcode']
        fieldnames = base_fieldnames + attr_columns

        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)

        return output.getvalue()

    def export_excel(self, include_variants=True, include_inactive=False):
        """Export products to Excel format"""
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl não instalado. Execute: pip install openpyxl")

        products = self.get_products(include_inactive)
        attr_columns = self._get_all_attr_columns(products) if include_variants else []

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Produtos"

        # Header
        base_headers = ['SKU', 'Nome', 'Tipo', 'Categoria', 'Marca', 'UOM', 'Estoque', 'Est. Mínimo', 'Custo', 'Código de Barras']
        attr_headers = [col.replace('attr_', '').title() for col in attr_columns]
        headers = base_headers + attr_headers

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)

        # Data rows
        row_num = 2
        for product in products:
            if product.product_type == ProductType.SIMPLE:
                row = self._simple_row(product)
                for col in attr_columns:
                    row[col] = ''
                self._write_excel_row(ws, row_num, row, attr_columns)
                row_num += 1
            else:
                row = self._parent_row(product)
                for col in attr_columns:
                    row[col] = ''
                self._write_excel_row(ws, row_num, row, attr_columns)
                row_num += 1

                if include_variants:
                    for variant in product.variants.filter(is_active=True):
                        row = self._variant_row(variant, attr_columns)
                        self._write_excel_row(ws, row_num, row, attr_columns)
                        row_num += 1

        # Auto-width columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _write_excel_row(self, ws, row_num, data, attr_columns):
        """Write a row to Excel worksheet"""
        values = [
            data['sku'],
            data['name'],
            data['type'],
            data['category'],
            data['brand'],
            data['uom'],
            data['stock'],
            data['minimum_stock'],
            data['cost'],
            data['barcode'],
        ]
        values.extend(data.get(col, '') for col in attr_columns)

        for col, value in enumerate(values, 1):
            ws.cell(row=row_num, column=col, value=value)

    def export_json(self, include_variants=True, include_inactive=False):
        """Export products to JSON format"""
        products = self.get_products(include_inactive)

        result = []
        for product in products:
            item = {
                'sku': product.sku,
                'name': product.name,
                'type': product.product_type,
                'category': product.category.name if product.category else None,
                'brand': product.brand.name if product.brand else None,
                'uom': product.uom,
                'description': product.description,
                'barcode': product.barcode,
                'is_active': product.is_active,
            }

            if product.product_type == ProductType.SIMPLE:
                item['stock'] = product.current_stock
                item['minimum_stock'] = product.minimum_stock
                item['cost'] = product.avg_unit_cost
            else:
                item['variants'] = []
                if include_variants:
                    for variant in product.variants.filter(is_active=True):
                        var_data = {
                            'sku': variant.sku,
                            'name': variant.name,
                            'barcode': variant.barcode,
                            'stock': variant.current_stock,
                            'minimum_stock': variant.minimum_stock,
                            'cost': variant.avg_unit_cost,
                            'attributes': {}
                        }
                        for attr_val in variant.attribute_values.all():
                            var_data['attributes'][attr_val.attribute_type.name] = attr_val.value
                        item['variants'].append(var_data)

            result.append(item)

        return json.dumps(result, indent=2, cls=DecimalEncoder, ensure_ascii=False)
